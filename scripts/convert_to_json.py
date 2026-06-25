# -*- coding: utf-8 -*-
"""
T1PKS/NRPS Biosynthesis Pathway Collection - Excel to JSON Converter

Converts the multi-sheet Excel file into a structured JSON database
suitable for web parsing/API consumption.

Data structure per manual:
- Each BGC (Biosynthetic Gene Cluster) has multiple ordered steps
- Steps reference previous products via Product_ID in Substrate field
- Nonlinearity captures domain skipping, iteration, module skipping, trans-AT
- Gene names use: ; for enzyme complexes, , for uncertain order, / for alternatives
"""

import json
import re
import openpyxl
from datetime import datetime
from collections import OrderedDict

INPUT_FILE = "data/raw/T1PKS, NRPS biosynthesis pathway collection Real final.xlsx"
OUTPUT_FILE = "data/normalized/biosynthesis_pathways.json"

HEADERS = [
    "MIBiG entry", "Compound name", "Class", "Order",
    "Enzyme", "Module", "Nonlinearity", "Substrate",
    "Product", "Product_ID", "Quality", "DOI"
]

DATA_SHEETS = ["지희", "병섭", "현우"]


def normalize_bgc_id(raw):
    """Normalize BGC ID to exactly BGC + 7 digits, e.g. BGC00000081 -> BGC0000081."""
    m = re.search(r'\d+', str(raw))
    return "BGC{:07d}".format(int(m.group())) if m else raw


def fix_datetime_product_id(val, bgc_number):
    """Fix Excel auto-date conversion: 1-1 -> Jan 1, 1-2 -> Jan 2, etc."""
    if isinstance(val, datetime):
        month = val.month
        day = val.day
        return f"{bgc_number}-{month}" if day == 0 else f"{bgc_number}-{day}" if month == 1 else f"{month}-{day}"
    return str(val) if val is not None else None


def fix_datetime_substrate(val, bgc_number):
    """Fix Excel auto-date conversion in Substrate column."""
    if isinstance(val, datetime):
        month = val.month
        day = val.day
        return f"{bgc_number}-{day}" if month == 1 else f"{month}-{day}"
    return str(val) if val is not None else None


def safe_str(val):
    if val is None:
        return None
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(val)
    if isinstance(val, datetime):
        return str(val)
    return str(val).strip()


def parse_nonlinearity(raw):
    """Parse the nonlinearity string into a structured dict."""
    if not raw:
        return None
    s = str(raw).strip()
    if not s or s == 'nan':
        return None
    # Return raw string - it's already in a quasi-dict format
    # that the frontend can parse or display
    return s


def parse_order(val):
    """Parse Order field which may contain branching markers like 5(1), 5(2)."""
    if val is None:
        return None
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(val)
    s = str(val).strip()
    if not s or s == 'nan':
        return None
    return s


def parse_module(val):
    if val is None:
        return None
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return str(val)
    s = str(val).strip()
    if not s or s == 'nan':
        return None
    return s


def extract_bgc_number(mibig_entry):
    """Extract the numeric part from BGC entry like BGC0000001 -> 1."""
    m = re.search(r'BGC0*(\d+)', str(mibig_entry))
    return int(m.group(1)) if m else None


def process_sheet(ws, sheet_name):
    """Process a single worksheet into a list of BGC pathway objects."""
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    
    bgcs = []
    current_bgc = None
    current_bgc_number = None
    
    for row_idx, row in enumerate(rows):
        # Pad short rows
        row = list(row) + [None] * (12 - len(row)) if len(row) < 12 else list(row[:12])
        
        mibig = normalize_bgc_id(safe_str(row[0])) if safe_str(row[0]) and safe_str(row[0]).startswith('BGC') else safe_str(row[0])
        compound = safe_str(row[1])
        cls = safe_str(row[2])
        order = row[3]
        enzyme = safe_str(row[4])
        module = row[5]
        nonlinearity = safe_str(row[6])
        substrate = row[7]
        product = safe_str(row[8])
        product_id = row[9]
        quality = safe_str(row[10])
        doi = safe_str(row[11])
        
        # New BGC entry
        if mibig and mibig.startswith('BGC'):
            if current_bgc:
                bgcs.append(current_bgc)
            
            current_bgc_number = extract_bgc_number(mibig)
            current_bgc = {
                "mibig_id": mibig,
                "compound_name": compound,
                "biosynthetic_class": cls,
                "doi": doi,
                "quality": quality if quality and quality.lower() not in ['nan'] else None,
                "source_sheet": sheet_name,
                "steps": []
            }
        elif mibig is None and current_bgc is not None:
            # Continuation row - may have compound name or class updates
            if compound and not current_bgc.get("compound_name"):
                current_bgc["compound_name"] = compound
            if cls and not current_bgc.get("biosynthetic_class"):
                current_bgc["biosynthetic_class"] = cls
            if doi and not current_bgc.get("doi"):
                current_bgc["doi"] = doi
            if quality and quality.lower() not in ['nan'] and not current_bgc.get("quality"):
                current_bgc["quality"] = quality
        
        # Skip rows with no meaningful step data
        if not enzyme and not product and product_id is None:
            continue
        
        if current_bgc is None:
            continue
        
        # Fix datetime issues
        fixed_product_id = fix_datetime_product_id(product_id, current_bgc_number) if current_bgc_number else safe_str(product_id)
        fixed_substrate = fix_datetime_substrate(substrate, current_bgc_number) if current_bgc_number else safe_str(substrate)
        
        # Parse substrate references
        substrate_str = fixed_substrate
        substrate_refs = []
        substrate_molecules = []
        if substrate_str:
            parts = substrate_str.split(';')
            for p in parts:
                p = p.strip()
                if re.match(r'^\d+-\d+', p) or re.match(r'^\d+$', p):
                    substrate_refs.append(p)
                else:
                    substrate_molecules.append(p)
        
        step = {
            "order": parse_order(order),
            "enzyme": enzyme,
            "module": parse_module(module),
            "nonlinearity": parse_nonlinearity(nonlinearity),
            "substrate": {
                "raw": substrate_str,
                "precursor_refs": substrate_refs if substrate_refs else None,
                "molecules": substrate_molecules if substrate_molecules else None
            },
            "product_smiles": product,
            "product_id": fixed_product_id,
        }
        
        current_bgc["steps"].append(step)
    
    if current_bgc:
        bgcs.append(current_bgc)
    
    return bgcs


def build_database():
    wb = openpyxl.load_workbook(INPUT_FILE, read_only=True)
    
    all_bgcs = []
    for sheet_name in DATA_SHEETS:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            bgcs = process_sheet(ws, sheet_name)
            all_bgcs.extend(bgcs)
            print(f"Sheet '{sheet_name}': {len(bgcs)} BGC pathways extracted")
    
    wb.close()
    
    # Build the final database structure
    database = {
        "metadata": {
            "title": "T1PKS/NRPS Biosynthesis Pathway Collection",
            "description": "Structured database of Type I PKS and NRPS biosynthetic pathways from MIBiG entries, curated from Actinomycetes-derived BGCs",
            "total_bgcs": len(all_bgcs),
            "total_steps": sum(len(b["steps"]) for b in all_bgcs),
            "source": "MIBiG (Minimum Information about a Biosynthetic Gene cluster)",
            "schema_version": "1.0",
            "fields": {
                "mibig_id": "MIBiG accession number (e.g. BGC0000001)",
                "compound_name": "Name of the biosynthesized compound",
                "biosynthetic_class": "PKS (Type I/Unknown/Iterative), NRPS, or hybrid; subclass in parentheses",
                "steps[].order": "Biosynthetic step order; (1),(2) suffix indicates branching",
                "steps[].enzyme": "Gene name(s): ; for complexes, , for uncertain order, / for alternatives, ? for unknown, Start for initial substrate",
                "steps[].module": "Module number (0=loading), TE for thioesterase release",
                "steps[].nonlinearity": "Domain skipping/inactive, iteration, module skipping, or trans-AT annotations",
                "steps[].substrate.raw": "Original substrate notation",
                "steps[].substrate.precursor_refs": "References to previous step Product_IDs",
                "steps[].substrate.molecules": "Named substrate molecules (e.g. Malonyl-CoA)",
                "steps[].product_smiles": "SMILES notation of the product (PPant S replaced with OH)",
                "steps[].product_id": "Unique identifier for this product (BGC#-step#)"
            }
        },
        "pathways": {b["mibig_id"]: b for b in all_bgcs}
    }
    
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        json.dump(database, f, ensure_ascii=False, indent=2)
    
    print(f"\nDatabase saved: {OUTPUT_FILE}")
    print(f"Total BGCs: {len(all_bgcs)}")
    print(f"Total steps: {sum(len(b['steps']) for b in all_bgcs)}")
    
    return database


if __name__ == "__main__":
    db = build_database()
